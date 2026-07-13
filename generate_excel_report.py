import sys
import subprocess
import os

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    install('openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "System Test Report"

# Data for the report based on our 3 rounds of testing
headers = [
    "STT", "Mã Test Case", "Tên Test Case", "Các bước thực hiện", "Dữ liệu đầu vào", 
    "Kết quả mong đợi", "Kết quả thực tế", "Trạng thái", "Ghi chú", 
    "Hình ảnh minh chứng", "Người thực hiện", "Ngày thực hiện", 
    "Môi trường", "Mức độ ưu tiên", "Loại Test"
]

data = [
    [1, "TC01", "Đăng nhập Admin", "1. Truy cập /auth/login\n2. Nhập Email & Password\n3. Nhấn Đăng nhập", "admin@digitallibrary.vn / admin123", "Đăng nhập thành công, vào Dashboard", "Đăng nhập thành công, Dashboard hiển thị đúng", "PASSED", "Đã xử lý chặn autofill của trình duyệt", "admin_dashboard_r2.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "High", "Functional"],
    [2, "TC02", "Tìm kiếm sách (Tiếng Việt)", "1. Tại thanh tìm kiếm, nhập 'Mắt biếc'\n2. Nhấn Enter", "Từ khóa: 'Mắt biếc'", "Hiển thị đúng sách 'Mắt biếc', không lỗi font", "Kết quả trả về chính xác, hiển thị đủ thông tin", "PASSED", "Đã sửa lỗi font UTF-8", "reader_search_success.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "High", "Functional"],
    [3, "TC03", "Hạn chế Role Guest", "1. Truy cập trực tiếp URL /admin/dashboard hoặc các trang admin mà không login", "URL: /admin/dashboard", "Redirect về trang login, không lộ data", "Hệ thống đẩy về trang login kèm thông báo", "PASSED", "Bảo mật phân quyền hoạt động tốt", "guest_redirect_login.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "Medium", "Security"],
    [4, "TC04", "Yêu cầu mượn sách (Reader)", "1. Đăng nhập Reader\n2. Vào trang sách, nhấn 'Mượn'\n3. Chọn ngày mượn/trả, nhấn gửi", "Reader: tranthib@example.com", "Yêu cầu được lưu, trạng thái 'pending'", "Yêu cầu #56 được tạo thành công, database ghi nhận 'pending'", "PASSED", "Luồng nghiệp vụ chuẩn", "reader_borrow_pending.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "High", "Functional"],
    [5, "TC05", "Duyệt yêu cầu mượn (Librarian)", "1. Đăng nhập Thủ thư\n2. Vào Quản lý mượn trả\n3. Nhấn 'Duyệt' cho yêu cầu #56", "Librarian: librarian@digitallibrary.vn", "Trạng thái chuyển 'Approved', gửi email báo", "Yêu cầu #56 được duyệt thành công, hiển thị đúng", "PASSED", "Cần click JS do UI conflict nhưng API chạy đúng", "librarian_approved_final.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "High", "Functional"],
    [6, "TC06", "Hiển thị Banner nổi bật", "1. Truy cập trang chủ", "N/A", "Banner hiển thị danh sách sách nổi bật dạng lưới", "Giao diện hiện đại, glassmorphism hiển thị đẹp", "PASSED", "Đã nâng cấp CSS & UI", "admin_dashboard.png", "Team 3", "2026-03-29", "Chrome/Tomcat 10", "Low", "UI"]
]

ws.append(headers)
for row in data:
    ws.append(row)

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4B5563") # Slate-600
alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.alignment = alignment
        cell.border = border
        if cell.column == 8: # Column H: Trạng thái
            if cell.value == "PASSED":
                cell.font = Font(color="10B981", bold=True) # Emerald-500
            else:
                cell.font = Font(color="EF4444", bold=True) # Red-500
        elif cell.column == 1: # STT
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Define specific column widths
widths = {
    1: 6,   # STT
    2: 15,  # Mã TC
    3: 30,  # Tên TC
    4: 40,  # Các bước
    5: 35,  # Đầu vào
    6: 35,  # Kết quả mong đợi
    7: 35,  # Kết quả thực tế
    8: 15,  # Trạng thái
    9: 30,  # Ghi chú
    10: 25, # Hình ảnh
    11: 15, # Người thực hiện
    12: 15, # Ngày
    13: 20, # Môi trường
    14: 15, # Mức độ
    15: 15  # Loại test
}

for col_num, width in widths.items():
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Freeze the top row
ws.freeze_panes = "A2"

# Save the file
output_path = r'C:\Users\tenma\Downloads\swp391-gr3-dung-merge-hao-borrow\Bao_Cao_Kiem_Thu_He_Thong.xlsx'
wb.save(output_path)
print(f"File saved successfully to: {output_path}")
